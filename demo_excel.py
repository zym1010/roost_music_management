# fetch all track info across all extracted JSONs,
# and then generate a huge excel file
# we can have two basic sheets
# 1. tracks
# 2. albums
# TODO
# 3. artists
# 4. album artists

# then we can have as many sheets as we like.
# basically smart playlists.
from collections import defaultdict
import json
from unicodedata import normalize
from os import path, walk
# tested with openpyxl 3.0.7
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import List
from itertools import groupby

PARENT_FOLDERS_TO_CHECK = [
    '/Users/yimengzh/Dropbox/music/output_20210806_with_mutagen/',
    '/Users/yimengzh/Dropbox/music/output_20210816_dsd/',
]


class Track:
    def __init__(self, *, file_path: str, subindex=-1, core_metadata: dict, extra_metadata: dict):
        self.file_path = file_path
        self.subindex = subindex
        self.core_metadata = core_metadata

        for k in self.core_metadata:
            if type(self.core_metadata[k]) is str:
                self.core_metadata[k] = normalize('NFC', self.core_metadata[k])

        self.extra_metadata = extra_metadata

        for k in self.extra_metadata:
            if type(self.extra_metadata[k]) is str:
                self.extra_metadata[k] = normalize('NFC', self.extra_metadata[k])

    @property
    def album_key(self):
        return self.core_metadata['ALBUM'], self.core_metadata['ALBUM_ARTIST']

    @property
    def track_key(self):
        return (self.file_path, self.subindex)


class TrackManager:
    def __init__(self):
        self.track_dict = dict()

    def add_track(self, track):
        assert track.track_key not in self.track_dict
        self.track_dict[track.track_key] = {
            'id': len(self.track_dict),
            'track': track,
        }

    def export_to_sheet(self, sheet: Worksheet):
        demo_track: Track = next(iter(self.track_dict.values()))['track']
        keys_core = sorted(demo_track.core_metadata.keys())
        keys_meta = sorted(demo_track.extra_metadata.keys())
        header = ['id'] + keys_core + keys_meta
        sheet.append(header)
        for key, v in self.track_dict.items():
            track = v['track']
            row_core = [
                track.core_metadata[x] for x in keys_core
            ]
            row_extra = [
                track.extra_metadata[x] for x in keys_meta
            ]
            row_this = [v['id']] + row_core + row_extra
            sheet.append(row_this)


class AlbumManager:
    def __init__(self):
        # disambiguate albums using raw album  name + album artist
        self.album_dict = dict()

    def add_track(self, track: Track):
        key = track.album_key

        if key not in self.album_dict:
            self.album_dict[key] = {
                'id': len(self.album_dict),
                'tracks': [],
            }

        self.album_dict[key]['tracks'].append(track)

    def check_album_consistency(self, tracks: List[Track]):
        try:
            disc_total_this = tracks[0].core_metadata['DISC_TOTAL']
            year_this = tracks[0].core_metadata['YEAR']
            genre_this = tracks[0].core_metadata['GENRE']

            tracks = sorted(tracks, key=lambda x: (x.core_metadata['DISC_NO'], x.core_metadata['TRACK_NO']))
            disc_no_all = []
            for disc_no, g in groupby(tracks, lambda x: x.core_metadata['DISC_NO']):
                g_to_use = list(g)
                assert 1 <= disc_no <= disc_total_this
                track_total_this = g_to_use[0].core_metadata['TRACK_TOTAL']

                for track_this in g_to_use:
                    assert track_this.core_metadata['TRACK_TOTAL'] == track_total_this
                    assert track_this.core_metadata['DISC_TOTAL'] == disc_total_this
                    assert track_this.core_metadata['YEAR'] == year_this, (track_this.core_metadata, year_this)
                    assert track_this.core_metadata['GENRE'] == genre_this
                assert [x.core_metadata['TRACK_NO'] for x in g_to_use] == list(range(1, track_total_this + 1))
                disc_no_all.append(disc_no)

            assert len(disc_no_all) == len(set(disc_no_all))
            assert set(disc_no_all) <= set(range(1, disc_total_this + 1))
            assert disc_no_all == list(range(1, disc_total_this + 1))
            return True
        except Exception:
            return False

    def export_to_sheet(self, sheet: Worksheet):
        sheet.append(
            ['id', 'ALBUM_ARTIST', 'ALBUM', 'TRACKS', 'CONSISTENT']
        )
        for key, v in self.album_dict.items():
            row_this = [v['id'], key[0], key[1], len(v['tracks'])]

            # check if it's well behaved.
            #
            row_this.append(
                self.check_album_consistency(v['tracks'])
            )

            sheet.append(row_this)


def get_all_libs():
    ret = []
    for parent_dir in PARENT_FOLDERS_TO_CHECK:
        for dirpath, dirnames, filenames in walk(parent_dir):
            if {'core_metadata', 'checksum', 'extra_metadata'} <= set(dirnames):
                ret.append(
                    path.join(parent_dir, dirpath)
                )

    return ret


def load_one_json(file_path):
    ret = dict()
    with open(file_path, 'rt', encoding='utf-8') as f:
        for line in f:
            line_json = json.loads(line)
            path_this = line_json['path_and_stat']['path']
            output_this = line_json['output']
            assert path_this not in ret
            ret[path_this] = output_this
    return ret


def parse_one_lib(
        lib_dir,
        track_manager: TrackManager,
        album_manager: AlbumManager,
):
    # get main.json from core and aux metadata
    # generate two dicts
    core_json = path.join(lib_dir, 'core_metadata', 'main.json')
    extra_json = path.join(lib_dir, 'extra_metadata', 'main.json')

    # load each file
    core_dict = load_one_json(core_json)
    extra_dict = load_one_json(extra_json)
    assert core_dict.keys() == extra_dict.keys()

    for file_path in core_dict:
        output_core = core_dict[file_path]
        output_extra = extra_dict[file_path]
        assert type(output_extra) is type(output_core)
        if type(output_core) is dict:
            # then just add
            track = Track(
                file_path=file_path,
                core_metadata=output_core,
                extra_metadata=output_extra,
            )

            track_manager.add_track(track)
            album_manager.add_track(track)
        else:
            assert type(output_core) is list
            assert len(output_core) == len(output_extra)
            for idx, (output_core_this, output_extra_this) in enumerate(zip(
                    output_core, output_extra
            )):
                track = Track(
                    file_path=file_path,
                    subindex=idx,
                    core_metadata=output_core_this,
                    extra_metadata=output_extra_this,
                )
                track_manager.add_track(track)
                album_manager.add_track(track)


def main():
    track_manager = TrackManager()
    album_manager = AlbumManager()
    output = 'demo_excel.xlsx'
    for lib_dir in get_all_libs():
        print(lib_dir)
        parse_one_lib(lib_dir, track_manager, album_manager)

    print(len(track_manager.track_dict), 'tracks')
    print(len(album_manager.album_dict), 'albums')

    # save
    wb = Workbook(write_only=True)
    # dump album
    album_manager.export_to_sheet(wb.create_sheet('albums'))
    # dump track
    track_manager.export_to_sheet(wb.create_sheet('tracks'))

    # save
    wb.save(output)


if __name__ == '__main__':
    main()
